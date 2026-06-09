"""IFTTT "Spotify -> spreadsheet" history ingestion (#76).

The owner's listening history arrives as a directory of ``.xlsx`` workbooks
written by the IFTTT applet that logs every Spotify play to a spreadsheet. Each
row is one play; there is **no header**. Columns, in order:

==  =============  =====================================
0   played_at      ``"December 3, 2024 at 12:31AM"``
1   track name     ``"Maniac"``
2   artist name    ``"Flower Face"``
3   spotify id     bare id ``"56O4WlGoUJiwfwoyRYqTe9"`` (no ``spotify:track:``)
4   spotify url    ``"https://open.spotify.com/track/56O4..."`` (ignored)
==  =============  =====================================

This is the only IFTTT-specific code; everything downstream sees source-agnostic
``ListenEvent``s. The bare spotify id maps straight onto ``TrackRef.spotify_id``
so its canonical id is ``spotify:<id>``.

TIMEZONE CAVEAT (flagged, not locked — see CONTEXT.md): IFTTT records the trigger
time as a *zone-less local wall-clock* string. V0 coerces it to UTC, which is
wrong by the local UTC offset and will skew temporal day-part/season buckets once
temporal roots run on real IFTTT data. It does **not** affect the honest-empty V0
run (no enrichment -> no temporal roots). Resolve the zone (or carry it through)
before trusting temporal roots derived from this source.
"""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from .models import ListenEvent, TrackRef
from .shared_store import canonical_track_id

SOURCE = "ifttt"
_MAX_SKIP_SAMPLES = 5
# "Month D, YYYY at H:MM<AM|PM>" — strptime is padding-lenient, so "3"/"03" and
# "2"/"02" both parse. %p matches the AM/PM with no separating space.
_TIMESTAMP_FORMAT = "%B %d, %Y at %I:%M%p"
_WORKBOOK_GLOB = "*.xlsx"


@dataclass
class IngestStats:
    """What ``load_ifttt_*`` dropped, so the CLI can surface it rather than let
    skips pass silently. A spike in ``skipped_unparseable`` is the loud signal
    that the export's timestamp shape drifted (vs aborting on the first odd row).
    """

    skipped_empty: int = 0
    skipped_no_identity: int = 0
    skipped_unparseable: int = 0
    unparseable_samples: list[str] = field(default_factory=list)

    def _note_unparseable(self, raw: str) -> None:
        self.skipped_unparseable += 1
        if len(self.unparseable_samples) < _MAX_SKIP_SAMPLES:
            self.unparseable_samples.append(raw)

    @property
    def total_skipped(self) -> int:
        return self.skipped_empty + self.skipped_no_identity + self.skipped_unparseable


def parse_ifttt_timestamp(raw: str) -> datetime:
    """Parse one IFTTT ``"Month D, YYYY at H:MM<AM|PM>"`` cell to a UTC-aware
    datetime. Raises ``ValueError`` (with the offending string) on any other
    shape — including the date-only ``"June 7, 2026"`` form seen once in the real
    export, which carries no time-of-day and so cannot be placed in a calendar
    bucket. Callers catch this and skip+count the row (see ``_row_to_event``)
    rather than fabricate a clock time the source never recorded.
    """
    return datetime.strptime(raw.strip(), _TIMESTAMP_FORMAT).replace(tzinfo=UTC)


def _cell(value: object) -> str:
    """Normalise a worksheet cell to a stripped string (``None`` -> ``""``)."""
    return "" if value is None else str(value).strip()


def _row_to_event(row: tuple, stats: IngestStats) -> ListenEvent | None:
    """Map one IFTTT row to a ``ListenEvent``; return ``None`` (and tally the
    reason on ``stats``) to skip it.

    Skipped: fully-blank/trailing rows, rows carrying neither a spotify id nor a
    name (no identity), and rows whose timestamp does not parse (unplaceable in
    time — never coerced to a fabricated clock time).
    """
    cells = [_cell(c) for c in row]
    while len(cells) < 4:  # tolerate short rows; columns 4+ (the url) are ignored
        cells.append("")
    played_raw, name, artist, spotify_id = cells[0], cells[1], cells[2], cells[3]

    if not played_raw and not spotify_id and not name:
        stats.skipped_empty += 1
        return None
    if not spotify_id and not name:
        stats.skipped_no_identity += 1
        return None
    try:
        played_at = parse_ifttt_timestamp(played_raw)
    except ValueError:
        stats._note_unparseable(played_raw)
        return None

    return ListenEvent(
        track=TrackRef(spotify_id=spotify_id or None, name=name, artist=artist),
        played_at=played_at,
        source=SOURCE,
    )


def load_ifttt_workbook(path: str | Path, *, stats: IngestStats | None = None) -> list[ListenEvent]:
    """Convert every data row of one IFTTT ``.xlsx`` to a ``ListenEvent``.
    Pass a shared ``stats`` to accumulate skip counts across many workbooks."""
    stats = stats if stats is not None else IngestStats()
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [
            event for row in ws.iter_rows(values_only=True) if (event := _row_to_event(row, stats))
        ]
    finally:
        wb.close()


def _dedup_key(event: ListenEvent) -> tuple[str, str]:
    """Identity for dedup: same track played at the same minute is one play.
    IFTTT timestamps are minute-resolution, so this is the finest honest grain."""
    return (canonical_track_id(event.track), event.played_at.isoformat())


def dedup_events(events: Iterable[ListenEvent]) -> list[ListenEvent]:
    """Drop duplicate plays (same canonical track id + played_at), keep first
    occurrence, then sort ascending by ``played_at`` for a stable history."""
    seen: set[tuple[str, str]] = set()
    out: list[ListenEvent] = []
    for event in events:
        key = _dedup_key(event)
        if key in seen:
            continue
        seen.add(key)
        out.append(event)
    out.sort(key=lambda e: e.played_at)
    return out


def load_ifttt_dir(
    directory: str | Path,
    *,
    pattern: str = _WORKBOOK_GLOB,
    stats: IngestStats | None = None,
) -> list[ListenEvent]:
    """Load and merge every IFTTT workbook under ``directory`` into one
    deduped, time-sorted history. Idempotent: re-running yields the same list.
    Pass a ``stats`` to receive skip counts across the whole directory."""
    root = Path(directory)
    stats = stats if stats is not None else IngestStats()
    events: list[ListenEvent] = []
    for workbook in sorted(root.glob(pattern)):
        events.extend(load_ifttt_workbook(workbook, stats=stats))
    return dedup_events(events)
